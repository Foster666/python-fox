import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from matplotlib import pyplot as plt, ticker


class SnObject:

    def __init__(self, recttime, machine, container, tottime, testtime, sernum, uuttype, area, result, test,
                 parentsernum, username, mode):
        self.recttime = str(recttime)[:10]
        self.machine = machine
        self.container = container
        self.tottime = tottime
        self.testtime = str(testtime)[11:16]
        self.sernum = sernum
        self.uuttype = uuttype
        self.area = area
        self.result = result
        self.test = str(test)[:6]
        self.parentsernum = parentsernum
        self.username = username
        self.mode = mode


class My_analyze_tool:
    def __init__(self, file_path, analyze_excel_name):
        self.file_path = file_path
        self.analyze_excel_name = analyze_excel_name
        self.sheet_name = "Sheet1"
        self.build_excel_name = "test_build.xlsx"
        self.data_container_analyze = True
        self.xlsx_num = 1
        self.xlsx_num_2 = 1

    def create_xlsx(self):
        try:
            try:
                os.remove(os.path.join(self.file_path, self.build_excel_name))
                print('刪除文件成功,重新創建')
            except:
                pass
            self.book = openpyxl.Workbook()
            self.sheet = self.book['Sheet']
            self.sheet2 = self.book.create_sheet('Sheet2')
            col_1 = self.sheet.column_dimensions[get_column_letter(1)]
            col_1.width = 20
            print(f'{self.build_excel_name}文件創建成功')
        except:
            print(f'{self.build_excel_name}文件創建失敗')
        self.xlsx_name = self.file_path + '\\' + self.build_excel_name
        return self.file_path + '\\' + self.build_excel_name

    def read_xslx_to_obj(self):
        book = openpyxl.load_workbook(f"{self.file_path}\\{self.analyze_excel_name}")
        sheet = book[self.sheet_name]
        result_objects_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            obj = SnObject(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
                           row[11], row[12])
            result_objects_list.append(obj)
        book.close()
        return result_objects_list

    def get_unique_value(self, obj_list=None, item_name=None):
        unique_value_list = []
        for obj in obj_list:
            value = getattr(obj, item_name)
            unique_value_list.append(value)
        return list(dict.fromkeys(unique_value_list))

    def find_server_and_container(self, result_obj):
        server_list = self.get_unique_value(obj_list=result_obj, item_name='machine')
        ser_con = {}
        for machine in server_list:
            machine_con = []
            for obj_con in result_obj:
                if obj_con.machine == machine:
                    machine_con.append(obj_con.container)
            machine_con = list(dict.fromkeys(machine_con))
            ser_con[machine] = machine_con
        return ser_con

    def get_top_fail(self, fail_list):
        top_fail = {}
        for sn_obj in fail_list:
            top_fail[sn_obj.test] = 0
        for sn_obj in fail_list:
            try:
                top_fail[sn_obj.test] += 1
            except:
                pass
        sorted_items = sorted(top_fail.items(), key=lambda x: x[1], reverse=True)
        top_fail = dict(sorted_items[:5])
        return top_fail

    def find_yield(self, ser_con=None, result_obj=None, datas_list=None):
        yield_result = {}
        for key, value in ser_con.items():
            yield_result[key] = {}
            for data in datas_list:
                yield_result[key][data] = {"all_list_mac": [], "pass_mac": [], "fail_mac": [], "top_fail_mac": {},
                                           "test_qty": 0, "1st_pass_qty": 0, "1st_fail_qty": 0, "1st_pass_yield": 0}
                for mac_sn_obj in result_obj:
                    if mac_sn_obj.recttime == data and mac_sn_obj.machine == key:
                        yield_result[key][data]["all_list_mac"].append(mac_sn_obj)
                        if mac_sn_obj.result == "P":
                            yield_result[key][data]["pass_mac"].append(mac_sn_obj)
                        else:
                            yield_result[key][data]["fail_mac"].append(mac_sn_obj)
                yield_result[key][data]['test_qty'] = f"{len(yield_result[key][data]['all_list_mac'])}"
                yield_result[key][data]['1st_pass_qty'] = f"{len(yield_result[key][data]['pass_mac'])}"
                yield_result[key][data]['1st_fail_qty'] = f"{len(yield_result[key][data]['fail_mac'])}"
                try:
                    yield_result[key][data][
                        '1st_pass_yield'] = f"{len(yield_result[key][data]['pass_mac']) / len(yield_result[key][data]['all_list_mac']) * 100:.2f}"
                except:
                    yield_result[key][data]['1st_pass_yield'] = '0.00'
                yield_result[key][data]["top_fail_mac"] = self.get_top_fail(yield_result[key][data]["fail_mac"])
            for container in value:
                yield_result[key][container] = {}
                for data in datas_list:
                    yield_result[key][container][data] = {"all_list": [], "pass": [], "fail": [], "top_fail": {},
                                                          "yield": ''}
                    for sn_obj in result_obj:
                        if sn_obj.machine == key and sn_obj.container == container and sn_obj.recttime == data:
                            yield_result[key][container][data]["all_list"].append(sn_obj)
                            if sn_obj.result == "P":
                                yield_result[key][container][data]["pass"].append(sn_obj)
                            else:
                                yield_result[key][container][data]["fail"].append(sn_obj)
                    try:
                        yield_result[key][container][data][
                            'yield'] = f"{len(yield_result[key][container][data]['pass']) / len(yield_result[key][container][data]['all_list']) * 100:.2f}"
                    except:
                        yield_result[key][container][data]['yield'] = '0.00'
                    yield_result[key][container][data]["top_fail"] = self.get_top_fail(
                        yield_result[key][container][data]["fail"])
        return yield_result, datas_list

    def draw_zhe(self, x_list=None, y_list=None, x_label=None, y_label=None, title=None, y_limit=None, png_name=None):
        if not y_limit:
            y_limit = 50
        png_name = png_name.replace(":", "_")
        y_float = [float(i) for i in y_list]
        x_list = [i[5:] for i in x_list]
        fig, ax = plt.subplots()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.plot(x_list, y_float, marker='o', color='red', linewidth=2)
        ax.set_xlabel(x_label, fontsize=12)
        ax.set_ylabel(y_label, fontsize=12)
        ax.set_title(title, y=1.05, fontsize=14)
        ax.set_ylim(bottom=y_limit, top=105)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.grid(True, axis='y', linestyle='--', alpha=0.7)
        for i in range(len(x_list)):
            plt.text(i, y_float[i] + 0.5, f"{str(y_float[i])}", ha="center")
        # plt.show()
        plt.savefig(f'{self.file_path}\\{png_name}.png', bbox_inches='tight')
        plt.close()
        print(f'已保存图 {png_name}')

    def draw_zhu(self, x_list=None, y_list=None, x_label=None, y_label=None, title=None, png_name=None):
        png_name = png_name.replace(":", "_")
        fig, ax = plt.subplots()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.bar(x_list, y_list)
        ax.set_xlabel(x_label, fontsize=12)
        ax.set_ylabel(y_label, fontsize=12)
        ax.set_title(title, y=1.05, fontsize=14)
        ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
        plt.xticks(rotation=10, ha='right')
        for i in range(len(x_list)):
            plt.text(i, y_list[i] + 0.1, str(y_list[i]), ha="center")
        # plt.show()
        plt.savefig(f'{self.file_path}\\{png_name}.png', bbox_inches='tight')
        plt.close()
        print(f'已保存{png_name}')

    def draw_all_fail_one_day(self, draw_message_one_day=None):
        x_list = list(draw_message_one_day['top_fail_dict'].keys())
        y_list = list(draw_message_one_day['top_fail_dict'].values())
        self.draw_zhu(x_list=x_list, y_list=y_list, x_label="fail item", y_label="数量",
                      title=f"{draw_message_one_day['machine']} {draw_message_one_day['container']} {draw_message_one_day['data'][5:]}",
                      png_name=f"{draw_message_one_day['machine']}_{draw_message_one_day['container']}_top_fail_{draw_message_one_day['data'][5:]}")

    def draw_analyze_container(self, ser_con=None, yield_result=None, datas_list=None, result_obj=None):
        self.top_fail_all_cell(ser_con=ser_con, result_obj=result_obj)  # container top fail 总的
        for key, value in ser_con.items():
            for container in value:
                y_data_rate = []
                for date_num in range(0, len(datas_list), 1):
                    data = datas_list[date_num]
                    y_data_rate.append(yield_result[key][container][data]["yield"])
                    draw_message_one_day = {'datas_list': datas_list,
                                            'top_fail_dict': yield_result[key][container][data]["top_fail"],
                                            'machine': key,
                                            'container': container,
                                            'data': data
                                            }
                    self.draw_all_fail_one_day(draw_message_one_day=draw_message_one_day)  # cell top fail 每天的
                self.draw_zhe(x_list=datas_list, y_list=y_data_rate, x_label="日期",
                              y_label="良率", title=f"{key} {container}", y_limit=50,
                              png_name=f"{key}_{container}_yield")  # cell 每天良率图

    def top_fail_all_foc(self, ser_con=None, result_obj=None):
        for key, value in ser_con.items():
            foc = {'all_list': [], 'pass_list': [], 'fail_list': [], 'top_fail': []}
            for sn_obj in result_obj:
                if sn_obj.machine == key:
                    foc['all_list'].append(sn_obj)
                    if sn_obj.result == 'P':
                        foc['pass_list'].append(sn_obj)
                    else:
                        foc['fail_list'].append(sn_obj)
            foc['top_fail'] = self.get_top_fail(foc['fail_list'])
            top_fail_item_name = x_list = list(foc['top_fail'].keys())
            y_list = list(foc['top_fail'].values())
            self.draw_zhu(x_list=x_list, y_list=y_list, x_label="fail item", y_label="数量",
                          title=f"{key} Top fail", png_name=f"{key}_top_fail")  # 服务器 top fail，总的
            for item_name in top_fail_item_name:
                item_num_dict = {}
                for con_obj in foc['fail_list']:
                    if con_obj.test == item_name:
                        container = con_obj.container[-5:]
                        if container in item_num_dict.keys():
                            item_num_dict[container] += 1
                        else:
                            item_num_dict[container] = 1
                x_list = list(item_num_dict.keys())
                y_list = list(item_num_dict.values())
                self.draw_zhu(x_list=x_list, y_list=y_list, x_label="UUT cell 号", y_label="数量",
                              title=f"{key} {item_name} distribute",
                              png_name=f"{key}_top_fail_{item_name}_distribute")  # 服务器 top fail 分布 总的

    def top_fail_all_cell(self, ser_con=None, result_obj=None):
        for key, value in ser_con.items():
            for container in value:
                cell = {'all_list': [], 'pass_list': [], 'fail_list': [], 'top_fail': []}
                for sn_obj in result_obj:
                    if sn_obj.machine == key and sn_obj.container == container:
                        cell['all_list'].append(sn_obj)
                        if sn_obj.result == 'P':
                            cell['pass_list'].append(sn_obj)
                        else:
                            cell['fail_list'].append(sn_obj)
                cell['top_fail'] = self.get_top_fail(cell['fail_list'])
                x_list = list(cell['top_fail'].keys())
                y_list = list(cell['top_fail'].values())
                self.draw_zhu(x_list=x_list, y_list=y_list, x_label="fail item", y_label="数量",
                              title=f"{key} {container} Top fail",
                              png_name=f"{key}_{container}_top_fail")  # cell top fail，总的

    def overall_analyze_foc(self, result_obj=None):
        datas_list = self.get_unique_value(obj_list=result_obj, item_name='recttime')  # 获得日期列表
        ser_con = self.find_server_and_container(result_obj=result_obj)  # 获得服务器和uut 列表
        self.top_fail_all_foc(ser_con=ser_con, result_obj=result_obj)  # 服务器 top fail 总的
        yield_result, datas_list = self.find_yield(ser_con=ser_con, result_obj=result_obj, datas_list=datas_list)
        for key, value_c in ser_con.items():
            y_foc_data_yield_list = []
            for data_num in range(0, len(datas_list), 1):
                data = datas_list[data_num]
                y_foc_data_yield_list.append(yield_result[key][data]['1st_pass_yield'])
                before = 3
                self.sheet.cell(self.xlsx_num, data_num + before).value = data[5:]
                self.sheet.cell(self.xlsx_num, before - 1).value = 'Date'
                self.sheet.cell(self.xlsx_num + 1, data_num + before).value = yield_result[key][data]['test_qty']
                self.sheet.cell(self.xlsx_num + 1, before - 1).value = 'Tested qty'
                self.sheet.cell(self.xlsx_num + 2, data_num + before).value = yield_result[key][data]['1st_pass_qty']
                self.sheet.cell(self.xlsx_num + 2, before - 1).value = '1st Passed qty'
                self.sheet.cell(self.xlsx_num + 3, data_num + before).value = yield_result[key][data]['1st_fail_qty']
                self.sheet.cell(self.xlsx_num + 3, before - 1).value = '1st Failed qty'
                self.sheet.cell(self.xlsx_num + 4, data_num + before).value = yield_result[key][data]['1st_pass_yield']
                self.sheet.cell(self.xlsx_num + 4, before - 1).value = '1st Passed yield'
                self.sheet.merge_cells(start_row=self.xlsx_num, start_column=1, end_row=self.xlsx_num + 4, end_column=1)
                merged_cell = self.sheet.cell(self.xlsx_num, 1)
                merged_cell.value = key
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                # print(
                #     f"{key} {data} test_qty {yield_result[key][data]['test_qty']} - {yield_result[key][data]['1st_pass_qty']} - "
                #     f"{yield_result[key][data]['1st_fail_qty']} - {yield_result[key][data]['1st_pass_yield']} - "
                #     f"{yield_result[key][data]['top_fail_mac']}"
                # )
                x_list = list(yield_result[key][data]['top_fail_mac'].keys())
                y_list = list(yield_result[key][data]['top_fail_mac'].values())
                self.draw_zhu(x_list=x_list, y_list=y_list, x_label="fail item", y_label="数量",
                              title=f"{key} {data} Top fail", png_name=f"{key}_{data}_top_fail")  # 服务器 top fail 每天的
            self.xlsx_num += 6
            self.draw_zhe(x_list=datas_list, y_list=y_foc_data_yield_list, x_label="日期",
                          y_label="良率", title=f"{key} yield", y_limit=50,
                          png_name=f"{key}_yield")  # 服务器 每天良率图
            for container_num in range(0, len(value_c), 1):
                container = value_c[container_num]
                for data_num in range(0, len(datas_list), 1):
                    data = datas_list[data_num]
                    # print(
                    #     f"{data} {key}, {container}, yield--> {yield_result[key][container][data]['yield']}, test--> {len(yield_result[key][container][data]['all_list'])}, "
                    #     f"pass--> {len(yield_result[key][container][data]['pass'])}, fail--> {len(yield_result[key][container][data]['fail'])}, "
                    #     f"Top fail--> {yield_result[key][container][data]['top_fail']}")
                    before = 3
                    self.sheet2.cell(self.xlsx_num_2, data_num + before).value = data[5:]
                    self.sheet2.cell(self.xlsx_num_2, before - 1).value = 'Date'
                    self.sheet2.cell(self.xlsx_num_2 + container_num + 1, data_num + before).value = \
                        yield_result[key][container][data]['yield']
                    self.sheet2.cell(self.xlsx_num_2 + container_num + 1, 2).value = container
                self.sheet2.merge_cells(start_row=self.xlsx_num_2, start_column=1,
                                        end_row=self.xlsx_num_2 + len(value_c), end_column=1)
                merged_cell = self.sheet2.cell(self.xlsx_num_2, 1)
                merged_cell.value = key
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            self.xlsx_num_2 += len(value_c) + 2
        return ser_con, yield_result, datas_list

    def main_run(self):
        xlsx_name = self.create_xlsx()
        result_obj = self.read_xslx_to_obj()
        ser_con, yield_result, datas_list = self.overall_analyze_foc(result_obj=result_obj)  # 按服务器分析
        self.draw_analyze_container(ser_con=ser_con, yield_result=yield_result, datas_list=datas_list,result_obj=result_obj) # 按cell分析
        self.book.save(self.xlsx_name)
        self.book.close()


file_path = r'E:\analyze_path'
analyze_excel_name = r'74-125084-01_PCBPM2_first_all_data_20230717_083216.xlsx'
j = My_analyze_tool(file_path=file_path, analyze_excel_name=analyze_excel_name)
j.main_run()
