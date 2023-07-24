import os
import re
import shutil
import sys
import json
import openpyxl
import xlwt
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import requests
import urllib3
import ssl  # 关闭安全请求警告
from urllib3.exceptions import InsecureRequestWarning

ssl._create_default_https_context = ssl._create_unverified_context
urllib3.disable_warnings(InsecureRequestWarning)


class SnObject:
    def __init__(self, recttime, machine, container, tottime, testtime, sernum, uuttype, area, result, test,
                 parentsernum, username, mode):
        self.recttime = recttime
        self.machine = machine
        self.container = container
        self.tottime = tottime
        self.testtime = testtime
        self.sernum = sernum
        self.uuttype = uuttype
        self.area = area
        self.result = result
        self.test = test
        self.parentsernum = parentsernum
        self.username = username
        self.mode = mode


class MainTool:

    def __init__(self):
        """ 初始化属性
        match = re.search(r'\w', string, flags=re.IGNORECASE)
        """
        self.file_path = 'test_log'
        self.baseurl = 'https://10.167.219.21/api/get-record-by-sernum/'
        self.headers = {'content-type': 'application/json'}
        self.p_base = 'https://10.167.219.21/api/get-test-log-list/?'
        self.p_host = '10.167.219.21'

    def read_txt(self, txt_path):
        """
        读取文本SN，以列表形式储存，列表会去重，SN以每行的形式存储
            传入txt_path 路径
        :param txt_path:
        :return: list(dict.fromkeys(read_list))
        """
        file_open = open(txt_path, "r", encoding="UTF-8-sig")
        read_list = []
        read_one_line = file_open.readline().strip()
        while read_one_line:
            read_list.append(read_one_line)
            read_one_line = file_open.readline().strip()
        return list(dict.fromkeys(read_list))

    def bulid_xls_by_xlwt(self, file_name):
        self.xlsx_name = f'{file_name}.xls'
        try:
            os.remove(self.xlsx_name)
            print('删除文件成功，重新创建')
        except:
            pass
        self.book = xlwt.Workbook(encoding='utf-8')
        self.sheet = self.book.add_sheet('Sheet', cell_overwrite_ok=True)
        self.sheet.col(0).width = 20 * 256
        self.sheet.write(1, 0, 'int(num)')
        self.sheet.write(0, 0, 'str(str1)')
        self.book.save(self.xlsx_name)

    def read_xslx(self, is_deduplication=None, path=None, sheet_name=None, max_col=None, min_col=None, min_row=None,
                  max_row=None):
        """
        读取SN，以列表形式储存，列表会去重
        :param path:
        :param sheet:
        :param max_col:
        :param min_col:
        :param min_row:
        :return:list(dict.fromkeys(read_list))
        """
        book = openpyxl.load_workbook(path)
        sheet = book[sheet_name]
        read_list = []
        if max_row:  # 有最大行数
            for one_row in sheet.iter_cols(max_col=max_col, min_col=min_col, min_row=min_row, max_row=max_row):
                for cell in one_row:
                    read_list.append(cell.value)
        else:  # 无最大行数
            for one_row in sheet.iter_cols(max_col=max_col, min_col=min_col, min_row=min_row):
                for cell in one_row:
                    read_list.append(cell.value)
        book.close()
        if is_deduplication:
            return list(dict.fromkeys(read_list))
        else:
            return read_list

    def read_tip(self, txt_path, tip=None):
        """
        :return: foc_list
        """
        while True:
            can_read = input(tip + '\t确认请输y\n')
            if can_read in ['y', 'Y']:
                try:
                    foc_list = self.read_txt(txt_path=txt_path)
                    len_all = len(foc_list)
                    if len_all > 5:
                        print('读取成功，输出前两个，和最后两个，仅供参考', foc_list[0], foc_list[1], foc_list[-2],
                              foc_list[-1])
                    elif len_all == 0:
                        print('读取到的个数为0')
                        read_exit = input('按任意按键结束程序\n')
                        sys.exit()
                    else:
                        print('读取成功，输出一个，仅供参考', foc_list[0])
                    return foc_list
                except:
                    print('您选择的读取格式为txt，请确定一行只有一个信息，保存格式为utf-8,读取期间请不要打开文件')
                    continue  # 继续小循环

    def print_message(self, foc_len=None, sn=None, key=None):
        """

        :param FOC:
        :param key:
        :return:
        """
        if key != foc_len - 1:
            print(f"成功保存{sn},已完成{key + 1}个,共{foc_len}")
        else:
            print(f"成功保存{sn},已完成{key + 1}个")
            print("程序完成退出")

    def create_folder(self, file_location=None):
        """

        :param file_location:
        :return: self.file_path
        """
        if file_location:
            self.file_path = file_location
        while True:
            commin_del_file = input(f"请确认当前路径下{self.file_path}文件夹内无重要文件，即将删除重建，确认请输入y\n")
            if commin_del_file in ['y', 'Y']:
                try:
                    if not os.path.exists(self.file_path):
                        os.makedirs(self.file_path)  # 创建文件夹
                        print('文件夹建立成功')
                    else:  # 否则就删除目录下的文件,再创建
                        shutil.rmtree(self.file_path, ignore_errors=True)
                        os.makedirs(self.file_path)  # 建立文件夹
                        print('文件夹建立成功')
                except:
                    print('创建文件夹失败,请重试,刪除创建时请误打开相同的文件夹,或文件夹内的文件')
                    continue  # 继续尝试创建
            else:
                continue  # 不是y，用户继续确认是否删除
            break  # 创建完成，跳出创建文件循环
        return self.file_path

    def input_station(self):
        """

        :return: PC_station
        """
        while True:
            try:
                PC_station = input(
                    '请選擇工站名，輸入數字 1:PCBDL\t2:PCBPM\t3:PCBPM2\t4:PCBPM3\n5:PCBPM4\t6:PCBASIC\t7:PCBTRX\t8:PCBST\t9:PCBP2\t10:other\n')
                if PC_station not in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10']:
                    print('工站选择错误')
                    continue
                if PC_station == '1':
                    PC_station = 'PCBDL'
                elif PC_station == '2':
                    PC_station = 'PCBPM'
                elif PC_station == '3':
                    PC_station = 'PCBPM2'
                elif PC_station == '4':
                    PC_station = 'PCBPM3'
                elif PC_station == '5':
                    PC_station = 'PCBPM4'
                elif PC_station == '6':
                    PC_station = 'PCBASIC'
                elif PC_station == '7':
                    PC_station = 'PCBTRX'
                elif PC_station == '8':
                    PC_station = 'PCBST'
                elif PC_station == '9':
                    PC_station = 'PCBP2'
                elif PC_station == '10':
                    PC_station = input('请输入工站名称').upper()
                return PC_station  # 选择工站成功
            except:
                print('请输入数字')
                continue

    def input_log_name(self):
        """

        :return: log_name
        """
        while True:
            try:
                log_name = input('请选择log输入数字\n 1:result\t2:console\t3:UUT\t4:SEQ\t5:other\n')
                if log_name not in ['1', '2', '3', '4', '5']:
                    print('log选择錯誤')
                    continue
                if log_name == '1':
                    log_name = 'result'
                elif log_name == '2':
                    log_name = 'console'
                elif log_name == '3':
                    log_name = 'UUT.log'
                elif log_name == '4':
                    log_name = 'SEQ'
                elif log_name == '5':
                    log_name = input('请输入日志名称所包含的特殊字段')
                return log_name
            except:
                print('请输入数字')
                continue

    def input_test_result(self, is_regular=True):
        """

        :return: return_test_result
        """
        if is_regular:
            while True:
                try:
                    test_result = input(
                        '请选择测试结果输入数字\n 1:最后一次pass log\t2:最后一次fail log\t3:所有pass log\n')
                    if test_result not in ['1', '2', '3']:
                        print('测试结果选择錯誤')
                        continue
                    return_test_result = {}
                    if test_result == '1':
                        return_test_result['test_result'] = ['P']
                        return_test_result['test_result_Explanation'] = 'last'  # 查询最后一次log
                    elif test_result == '2':
                        return_test_result['test_result'] = ['F']
                        return_test_result['test_result_Explanation'] = 'last'  # 查询最后一次log
                    elif test_result == '3':
                        return_test_result['test_result'] = ['P']
                        return_test_result['test_result_Explanation'] = 'all'  # 查询最后一次log
                    return return_test_result
                except:
                    print('请输入数字')
                    continue
        else:
            while True:
                try:
                    test_result = input(
                        '请选要下载的测试结果输入数字\n 1:最后一次pass log\t2:最后一次fail log\t3:所有的pass的log\t4:所有的fail的log\t5:所有的pass和fail log\t6:第一次pass log\t7:第一次fail log\n')
                    if test_result not in ['1', '2', '3', '4', '5', '6', '7']:
                        print('测试结果选择錯誤')
                        continue
                    return_test_result = {}
                    if test_result == '1':
                        return_test_result['test_result'] = ['P']
                        return_test_result['test_result_Explanation'] = 'last'  # 下载最后一次log
                    elif test_result == '2':
                        return_test_result['test_result'] = ['F']
                        return_test_result['test_result_Explanation'] = 'last'  # 下载最后一次log
                    elif test_result == '3':
                        return_test_result['test_result'] = ['P']
                        return_test_result['test_result_Explanation'] = 'all'  # 下载所有log
                    elif test_result == '4':
                        return_test_result['test_result'] = ['F']
                        return_test_result['test_result_Explanation'] = 'all'  # 下载所有log
                    elif test_result == '5':
                        return_test_result['test_result'] = ['P', 'F']
                        return_test_result['test_result_Explanation'] = 'all'  # 下载所有log
                    if test_result == '6':
                        return_test_result['test_result'] = ['P']
                        return_test_result['test_result_Explanation'] = 'first'  # 下载第一次log
                    elif test_result == '7':
                        return_test_result['test_result'] = ['F']
                        return_test_result['test_result_Explanation'] = 'first'  # 下载第一次log
                    return return_test_result
                except:
                    print('请输入数字')
                    continue

    def input_log_month(self):
        """

        :return: month_input
        """

        while True:
            try:
                month_input = input(
                    '请选择查询最近的月数\n1:最近1个月\t2:最近3个月\t3:最近6个月\t4:最近12个月\t5:最近24个月\n')
                if month_input not in ['1', '2', '3', '4', '5']:
                    print('查询最近月数选择錯誤')
                    continue
                if month_input == '1':
                    month_input = 1
                elif month_input == '2':
                    month_input = 3
                elif month_input == '3':
                    month_input = 6
                elif month_input == '4':
                    month_input = 12
                elif month_input == '5':
                    month_input = 24
                return month_input
            except:
                print('请输入数字')
                continue

    def selection_parameters(self, station=None, log_name=None, test_result=None, month=None, is_regular=None):
        """

        :param station:
        :param log_name:
        :param test_result:
        :param month:
        :return: self.selection_parameters_dict
        """
        self.selection_parameters_dict = {}
        if station:  # 表示有工站名称输入
            self.selection_parameters_dict['station'] = self.input_station()
        if log_name:  # 表示有log名称输入
            self.selection_parameters_dict['log_name'] = self.input_log_name()
        if test_result:  # 表示有test_result输入
            self.selection_parameters_dict['test_result'] = self.input_test_result(is_regular)
        if month:  # 表示有月份输入
            self.selection_parameters_dict['month'] = self.input_log_month()
        return self.selection_parameters_dict

    def requests_alldate_from_sn_and_month(self, sn=None):
        """

        :param sn:
        :return: self.data_list
        """
        data = {"sernum": sn, "month": self.selection_parameters_dict['month']}
        data_list = requests.post(url=self.baseurl, data=json.dumps(data), headers=self.headers, verify=False).json()[
            'data']
        data_list.reverse()
        # data_list.encoding = 'utf-8'
        # data_list = json.loads(data_list.text)
        self.data_list = data_list  # 所有的测试记录翻转,按照浏览器查询的结果排列
        return self.data_list

    def filter_area(self, data_list, area):
        filter_area_list = []
        for i_num_data in range(0, len(data_list), 1):
            if data_list[i_num_data]['area'] == area:
                filter_area_list.append(data_list[i_num_data])
        return filter_area_list

    def filter_result(self, data_list, result):
        filter_result_list = []
        for i_num_data in range(0, len(data_list), 1):
            if data_list[i_num_data]['result'] in result:
                filter_result_list.append(data_list[i_num_data])
        return filter_result_list

    def filter_mode(self, data_list, mode):
        filter_result_list = []
        for i_num_data in range(0, len(data_list), 1):
            if data_list[i_num_data]['mode'] == mode:
                filter_result_list.append(data_list[i_num_data])
        return filter_result_list

    def filter_test(self, data_list, test):
        filter_result_list = []
        for i_num_data in range(0, len(data_list), 1):
            if data_list[i_num_data]['test'] == test:
                filter_result_list.append(data_list[i_num_data])
        return filter_result_list

    def filter_data(self, record_time=None, test=None, serial_number=None, uut_type=None, area=None, mode=None,
                    result=None, fail_name=None, test_time=None, machine=None, container=None):
        """
        :return:
        """
        """
        {'id': 121190130,'machine': 'focwnbu188','container': 'PCBDL:UUT02','recttime': '2022-11-16T16:18:39','tottime': '00:23:25', 
        'testtime': '2022-11-16 23:55:13.746993+08:00','sernum': 'FOC26466TM3','uuttype': '74-127115-02','area': 'PCBDL','result': 'P', 
        'test': （當是F的時候）'74_TX_5310_ANT7_EVM_-34_22','partnum': '','partnumrev': '','tannum': '', 'tannumrev': '','pid': '','vid': '', 
        'parentsernum': '','username': 'genius','mode': 'PROD','order': '','lineid': '', 'bflush': 0, 'bitmap': 0,'swrev': '','diagrev': '', 
        'hwrev': '','deviation': '','label': '','license': '','str1name': 'MB_MAC','str1': '1C:FC:17:2B:08:64','str2name': '2G_MAC', 
        'str2': '10:A8:29:AC:67:50','str3name': '5G_MAC','str3': '10:A8:29:AC:67:60','str4name': '6G_MAC','str4': '10:A8:29:AC:67:70', 
        'str5name': '', 'str5': '','str6name': '','str6': '','sync': 1}
        """
        filter_list = self.data_list
        if area:
            filter_list = self.filter_area(filter_list, area)
        if result:
            filter_list = self.filter_result(filter_list, result)
        if mode:
            filter_list = self.filter_mode(filter_list, mode)
        if test:
            filter_list = self.filter_test(filter_list, test)
        return filter_list

    def get_test_log_list(self, filter_list):
        test_log_list = []
        for i_num_data in range(0, len(filter_list), 1):
            timestamp = filter_list[i_num_data]['testtime']
            machine = filter_list[i_num_data]['machine']
            container = filter_list[i_num_data]['container']
            test_log_list_url = self.p_base + 'time_stamp=' + timestamp.replace(' ',
                                                                                '%20') + '&machine=' + machine + '&container=' + container + '&host=' + self.p_host
            test_log_list_url_data = \
                requests.get(url=test_log_list_url, headers=self.headers, verify=False).json()['payload']['data']
            test_log_list_url_data.reverse()
            for i_num in range(0, len(test_log_list_url_data), 1):
                test_log_list.append(test_log_list_url_data[i_num])
        return test_log_list

    def get_all_log_url(self, test_log_list, log_name=None):
        log_url_list = []
        for i_num_url in range(0, len(test_log_list), 1):
            try:
                test_log_list[i_num_url]['name'].index(log_name)
                log_url = test_log_list[i_num_url]['url']  # 找到log_name的url
                log_url_list.append(log_url)
            except:
                pass
        return log_url_list

    def get_all_log(self, log_url):
        log_text_list = []
        for i_num in range(0, len(log_url), 1):
            if '.log' in log_url[i_num]:
                log_text = requests.get(url=log_url[i_num], headers=self.headers, verify=False).text
                log_text_list.append(log_text)
            elif '.txt' in log_url[i_num]:
                log_text = requests.get(url=log_url[i_num], headers=self.headers, verify=False).text
                log_text_list.append(log_text)
            else:
                log_text = requests.get(url=log_url[i_num], headers=self.headers, verify=False).content
                log_text_list.append(log_text)
        return log_text_list

    def save_log(self, sn=None, log_name=None, log_text=None):
        save_file_name = sn + '~' + log_name.split('/')[-1].replace(':', '_').replace('.log', '.txt')
        newfile = self.file_path + '\\' + save_file_name
        if '.txt' in newfile:
            with open(newfile, 'w', encoding="utf-8") as new_file:  # wb(只写+二进制模式) wb+(读写) 从头开始 ab(追加写) ab+(追加读写)
                new_file.write(log_text)
                new_file.close()
        else:
            with open(newfile, 'wb') as new_file2:
                new_file2.write(log_text)
                new_file2.close()

    def save_sn(self, name=None, sn=None):
        if name:
            pass
        else:
            name = 'expect_sn'
        save_sn_name = self.file_path + '\\' + name + '.txt'
        with open(save_sn_name, 'a', encoding="utf-8") as save_sn_name:
            save_sn_name.write(sn + '\n')
            save_sn_name.close()

    def create_xlsx(self, is_initial=False, xlsx_name=None):
        if xlsx_name:
            xlsx_name = xlsx_name + '.xlsx'
        else:
            xlsx_name = 'test_build.xlsx'
        if is_initial:
            try:
                try:
                    os.remove(self.file_path + '\\' + xlsx_name)
                    print('刪除文件成功,重新創建')
                except:
                    pass
                book = openpyxl.Workbook()
                sheet = book['Sheet']
                col_1 = sheet.column_dimensions[get_column_letter(1)]
                col_1.width = 20
                col_2 = sheet.column_dimensions[get_column_letter(2)]
                col_2.width = 20
                col_3 = sheet.column_dimensions[get_column_letter(3)]
                col_3.width = 20
                col_4 = sheet.column_dimensions[get_column_letter(4)]
                col_4.width = 20
                col_5 = sheet.column_dimensions[get_column_letter(5)]
                col_5.width = 20
                col_6 = sheet.column_dimensions[get_column_letter(6)]
                col_6.width = 20
                col_7 = sheet.column_dimensions[get_column_letter(7)]
                col_7.width = 40
                col_8 = sheet.column_dimensions[get_column_letter(8)]
                col_8.width = 20
                col_9 = sheet.column_dimensions[get_column_letter(9)]
                col_9.width = 20
                col_10 = sheet.column_dimensions[get_column_letter(10)]
                col_10.width = 20
                sheet.cell(1, 1).value = 'Record Time'
                sheet.cell(1, 2).value = 'Serial Number'
                sheet.cell(1, 3).value = 'UUT Type'
                sheet.cell(1, 4).value = 'Area'
                sheet.cell(1, 5).value = 'Mode'
                sheet.cell(1, 6).value = 'Result'
                sheet.cell(1, 7).value = 'Fail Name'
                sheet.cell(1, 8).value = 'Test time'
                sheet.cell(1, 9).value = 'Machine'
                sheet.cell(1, 10).value = 'Container'
                book.save(self.file_path + '\\' + xlsx_name)
                book.close()
                print('文件創建成功')
            except:
                print(f'{xlsx_name}文件創建失敗')
        else:
            try:
                try:
                    os.remove(self.file_path + '\\' + xlsx_name)
                    print('刪除文件成功,重新創建')
                except:
                    pass
                book = openpyxl.Workbook()
                sheet = book['Sheet']
                col_1 = sheet.column_dimensions[get_column_letter(1)]
                col_1.width = 20
                book.save(self.file_path + '\\' + xlsx_name)
                book.close()
                print('文件創建成功')
            except:
                print(f'{xlsx_name}文件創建失敗')
        return self.file_path + '\\' + xlsx_name

    def find_evm_num(self, filter_list):
        evm_list = []
        for i_num in range(0, len(filter_list), 1):
            if 'EVM' in filter_list[i_num]['test']:
                evm_list.append(filter_list[i_num]['test'].split('_')[0])
        return list(dict.fromkeys(evm_list))

    def found_evm_value(self, evm_list):
        evm_list_value = []
        result_log = ''
        filter_list = self.filter_data(area=self.selection_parameters_dict['station'], result='P')  # 添加筛选条件
        test_log_list = self.get_test_log_list(filter_list)
        if test_log_list:  # 有log
            try:
                log_url_list = [self.get_all_log_url(test_log_list, log_name='result')[0]]
            except:
                log_url_list = [self.get_all_log_url(test_log_list, log_name='Output')[0]]
            if log_url_list:
                # print('有log，且有对应log_name')
                result_log = self.get_all_log(log_url_list)[0]
            else:
                # print('有log，但是无对应log name')
                self.save_sn(sn=filter_list[0]['sernum'])
            for get_num in range(0, len(evm_list), 1):
                try:
                    res_evm_num_value_all = re.findall('EVM[\s\S]*?Freq', re.search(
                        f'\n{evm_list[get_num]}. TX_VERIFY    ath[\s\S]*?Test Result', result_log).group())
                    for i_retry_num in range(0, len(res_evm_num_value_all),
                                             1):  # retry多少次就會循環多少次,因為每次retry都會匹配到evm一次findall
                        if 'x' not in res_evm_num_value_all[i_retry_num]:
                            res_evm_num_value = re.search(r'-[\s\S]* ', re.search('EVM[\s\S]*?dB',
                                                                                  res_evm_num_value_all[
                                                                                      i_retry_num]).group()).group().strip()
                            evm_list_value.append(res_evm_num_value)
                except:
                    self.save_sn(sn=filter_list[0]['sernum'])  # print('找evm失败')
        else:
            pass  # print('evm fail 找不到pass result')
        return evm_list_value

    def combine_evm(self, evm_list=None, evm_list_value=None):
        p_test = ''
        for i_num in range(0, len(evm_list_value), 1):
            p_test = p_test + f"{evm_list[i_num]}_EVM的值：{evm_list_value[i_num]},"
        return p_test

    def regular_log(self, all_log_list=None, re_list=None):
        has_such_list = []
        is_pass = False
        try:
            for i_re_num in range(0, len(re_list), 1):
                if is_pass:
                    is_pass = False
                else:
                    if re_list[i_re_num] in ['regular_2:', 'regular_3:', 'regular_4:', 'regular_5:', 'regular_6:']:
                        has_such_second = re.search((re_list[i_re_num + 1]), has_such_list[-1]).group()
                        has_such_list.append(has_such_second)
                        is_pass = True
                    else:
                        has_such_txt = re.search(re_list[i_re_num], all_log_list).group()
                        has_such_list.append(has_such_txt)
        except:
            return has_such_list
        return has_such_list

    def merge_xlsx_cell(self, xlsx_name, cell_str, cell_value):
        """
        wb = Workbook()
        ws = wb.active
        ws.merge_cells(range_string='A1:B3')
        ws.merge_cells(start_row=5, start_column=4, end_row=8, end_column=8)
        """
        wb = openpyxl.load_workbook(filename=xlsx_name)
        sheet = wb.active
        sheet.merge_cells(range_string=cell_str)
        cell_start = cell_str.split(':')[0]
        sheet[cell_start] = cell_value
        wb.save(xlsx_name)

    def unmerge_xlsx_cell(self, xlsx_name, cell_str):
        wb = openpyxl.load_workbook(filename=xlsx_name)
        sheet = wb.active
        sheet.unmerge_cells(range_string=cell_str)
        wb.save(xlsx_name)

    def sn_find_mac_addr(self):
        sn_list = []
        for key in range(0, len(sn_list), 1):
            macurl = 'https://10.167.219.21/api/genius/'
            headers = {'content-type': 'application/json'}
            data = {"data": {"mac": sn_list[key]}, "service": "get-mac-record"}
            try:
                mac_addr = \
                requests.post(url=macurl, data=json.dumps(data), headers=headers, verify=False).json()['payload'][
                    'data'][0]['mac']
            except:
                mac_addr = ''
            print(mac_addr)

    def down_log(self):
        self.create_folder()
        self.selection_parameters(station=1, log_name=1, month=1, test_result=1)
        foc_list = self.read_tip(txt_path='SN.txt', tip='请把sn保存在当前文件夹内，命名为SN.txt')
        for key in range(0, len(foc_list), 1):
            self.requests_alldate_from_sn_and_month(foc_list[key])
            filter_list = self.filter_data(area=self.selection_parameters_dict['station'],  # 添加筛选条件
                                           result=self.selection_parameters_dict['test_result']['test_result'],
                                           mode='PROD',
                                           test='PING_TEST_FAIL')
            test_log_list = self.get_test_log_list(filter_list)
            if test_log_list:
                # print('有log')
                log_url_list = self.get_all_log_url(test_log_list, log_name=self.selection_parameters_dict['log_name'])
                if log_url_list:
                    # print('有log，且有对应log_name')
                    if self.selection_parameters_dict['test_result']['test_result_Explanation'] == 'last':
                        log_url_list = [log_url_list[0]]
                    elif self.selection_parameters_dict['test_result']['test_result_Explanation'] == 'first':
                        log_url_list = [log_url_list[-1]]
                else:
                    # print('有log，但是无对应log name')
                    self.save_sn(sn=foc_list[key])
                all_log_list = self.get_all_log(log_url_list)
                for i_num in range(0, len(all_log_list), 1):
                    self.save_log(sn=foc_list[key], log_name=log_url_list[i_num], log_text=all_log_list[i_num])
                self.print_message(foc_len=len(foc_list), sn=foc_list[key], key=key)
            else:
                # print('无log')
                self.save_sn(sn=foc_list[key])

    def check_test_result(self):
        self.create_folder()
        self.selection_parameters(station=1, month=1)
        foc_list = self.read_tip(txt_path='SN.txt', tip='请把sn保存在当前文件夹内，命名为SN.txt')
        xlsx_path = self.create_xlsx(is_initial=True)
        one_record = 0
        p_test = ''
        book = openpyxl.load_workbook(xlsx_path)
        sheet = book['Sheet']
        for key in range(0, len(foc_list), 1):
            self.requests_alldate_from_sn_and_month(foc_list[key])
            filter_list = self.filter_data(area=self.selection_parameters_dict['station'],  # 添加筛选条件
                                           )
            evm_list = False
            try:
                evm_list = self.find_evm_num(filter_list)
            except:
                pass
            if evm_list:
                evm_list_value = self.found_evm_value(evm_list)
                p_test = self.combine_evm(evm_list=evm_list, evm_list_value=evm_list_value)
            else:
                pass  # print(f'直接写入记录{foc_list[key]}')
            for i_num in range(0, len(filter_list), 1):
                if filter_list[i_num]['result'] != 'S':
                    ff = key + 2 + one_record  # 加空格
                    # ff = 2 + one_record   # 不加空格
                    one_record += 1
                    sheet.cell(ff, 1).value = filter_list[i_num]['recttime'].replace('T', ' ')
                    sheet.cell(ff, 2).value = filter_list[i_num]['sernum']
                    sheet.cell(ff, 3).value = filter_list[i_num]['uuttype']
                    sheet.cell(ff, 4).value = filter_list[i_num]['area']
                    sheet.cell(ff, 5).value = filter_list[i_num]['mode']
                    sheet.cell(ff, 6).value = filter_list[i_num]['result']
                    if filter_list[i_num]['result'] == 'P':
                        p_test_insert = p_test
                        sheet.cell(ff, 7).value = p_test_insert
                        p_test = ''
                        # sheet.cell(ff, 7).fill = PatternFill(start_color="009933", fill_type="solid")
                    else:
                        sheet.cell(ff, 7).value = filter_list[i_num]['test']
                    sheet.cell(ff, 8).value = filter_list[i_num]['tottime']
                    sheet.cell(ff, 9).value = filter_list[i_num]['machine']
                    sheet.cell(ff, 10).value = filter_list[i_num]['container']
            self.print_message(foc_len=len(foc_list), sn=foc_list[key], key=key)
        book.save(xlsx_path)
        book.close()

    def regular_text(self):
        self.create_folder()
        self.selection_parameters(station=1, log_name=1, month=1, test_result=1, is_regular=True)
        foc_list = self.read_tip(txt_path='SN.txt', tip='请把sn保存在当前文件夹内，命名为SN.txt')
        re_list = self.read_tip(txt_path='str.txt',
                                tip='请把匹配信息保存在当前路径并命名为str.txt。保持一行一个条件，可保存多个')
        xlsx_path = self.create_xlsx(is_initial=False)
        book = openpyxl.load_workbook(xlsx_path)
        sheet = book['Sheet']
        sheet.cell(1, 1).value = 'no match any thing SN'
        sheet.cell(1, 2).value = 'SN have match'
        col_2 = sheet.column_dimensions[get_column_letter(2)]
        col_2.width = 20
        col_3 = sheet.column_dimensions[get_column_letter(3)]
        col_3.width = 20
        col_4 = sheet.column_dimensions[get_column_letter(4)]
        col_4.width = 20
        book.save(xlsx_path)
        book.close()
        one_record = 2
        second_record = 2
        for key in range(0, len(foc_list), 1):
            self.requests_alldate_from_sn_and_month(foc_list[key])
            filter_list = self.filter_data(area=self.selection_parameters_dict['station'],  # 添加筛选条件
                                           result=self.selection_parameters_dict['test_result']['test_result'])
            test_log_list = self.get_test_log_list(filter_list)
            if test_log_list:  # 有log
                # print('有log')
                log_url_list = self.get_all_log_url(test_log_list, log_name=self.selection_parameters_dict['log_name'])
                if log_url_list:  # 有对应log_name地log
                    # print('有log，且有对应log_name')
                    if self.selection_parameters_dict['test_result']['test_result_Explanation'] == 'last':
                        log_url_list = [log_url_list[0]]
                else:
                    # print('有log，但是无对应log name')
                    self.save_sn(sn=foc_list[key])
                all_log_list = self.get_all_log(log_url_list)
                for i_num in range(0, len(all_log_list), 1):
                    such_txt = self.regular_log(all_log_list[i_num], re_list)
                    if such_txt:
                        book = openpyxl.load_workbook(xlsx_path)
                        sheet = book['Sheet']
                        for one_row in sheet.iter_cols(max_col=2, min_col=2, min_row=(one_record - 1),
                                                       max_row=(one_record - 1)):
                            for cell in one_row:
                                if cell.value == filter_list[0]['sernum']:
                                    pass
                                else:
                                    ff = one_record
                                    one_record += 1
                                    sheet.cell(ff, 2).value = filter_list[0]['sernum']
                                    for i_num_such in range(0, len(such_txt), 1):
                                        text = ILLEGAL_CHARACTERS_RE.sub(r'', such_txt[i_num_such]).replace(' ',
                                                                                                            '')  # 最后去除空格
                                        sheet.cell(ff, 3 + i_num_such).value = text
                                    book.save(xlsx_path)
                                    book.close()
                    else:
                        book = openpyxl.load_workbook(xlsx_path)
                        sheet = book['Sheet']
                        for one_row in sheet.iter_cols(max_col=1, min_col=1, min_row=(second_record - 1),
                                                       max_row=(second_record - 1)):
                            for cell in one_row:
                                if cell.value == filter_list[0]['sernum']:
                                    pass
                                else:
                                    ff = second_record
                                    second_record += 1
                                    sheet.cell(ff, 1).value = filter_list[0]['sernum']
                                    book.save(xlsx_path)
                                    book.close()
            else:
                # print('无log')
                self.save_sn(sn=foc_list[key], name='sn_no_log')
            self.print_message(foc_len=len(foc_list), sn=foc_list[key], key=key)

    def find_machine_num_and_uut(self, machine=None):
        num_list = []
        for key_num in range(0, len(self.container_list), 1):
            if machine == self.machine_list[key_num]:
                self.test_result[machine]['uut'].setdefault(self.container_list[key_num], {})
                num_list.append(key_num)
        return num_list

    def creat_dict(self, item=None, machine=None):
        for uut in self.test_result[machine]['uut']:
            self.test_result[machine]['uut'][uut][item] = {}
            self.test_result[machine]['uut'][uut][item]['num'] = []
            self.test_result[machine]['uut'][uut][item]['amount'] = 0
            self.test_result[machine]['uut'][uut][item]['num_list'] = []

    def find_result(self, item=None, machine=None):
        for key in self.test_result[machine]['num']:
            for gets in self.test_result[machine]['uut']:
                if gets == self.container_list[key]:
                    if self.test_list[key] is not None:
                        if item in self.test_list[key]:
                            self.test_result[machine]['uut'][gets][item]['num'].append(key)
                            ant_num = re.search('ANT\d', self.test_list[key]).group()
                            try:
                                self.test_result[machine]['uut'][gets][item][ant_num] += 1
                                self.test_result[machine]['uut'][gets][item]['amount'] += 1
                                self.test_result[machine]['uut'][gets][item]['num_list'].append(ant_num)
                                self.test_result[machine]['uut'][gets][item]['num_list'] = list(
                                    dict.fromkeys(self.test_result[machine]['uut'][gets][item]['num_list']))
                            except:
                                self.test_result[machine]['uut'][gets][item][ant_num] = 1
                                self.test_result[machine]['uut'][gets][item]['amount'] += 1
                                self.test_result[machine]['uut'][gets][item]['num_list'].append(ant_num)
                                self.test_result[machine]['uut'][gets][item]['num_list'] = list(
                                    dict.fromkeys(self.test_result[machine]['uut'][gets][item]['num_list']))

    def count_fail(self):
        while True:
            can_read = input('请把下载的xlsx文件命名成 data.xlsx  确认请输y\n')
            if can_read in ['y', 'Y']:
                break
            else:
                continue
        print('正在读取文件')
        self.machine_list = self.read_xslx(is_deduplication=False, path='data.xlsx', sheet_name='Sheet1', max_col=2,
                                           min_col=2, min_row=2)
        self.container_list = self.read_xslx(is_deduplication=False, path='data.xlsx', sheet_name='Sheet1', max_col=3,
                                             min_col=3, min_row=2)
        self.result_list = self.read_xslx(is_deduplication=False, path='data.xlsx', sheet_name='Sheet1', max_col=9,
                                          min_col=9, min_row=2)
        self.test_list = self.read_xslx(is_deduplication=False, path='data.xlsx', sheet_name='Sheet1', max_col=10,
                                        min_col=10, min_row=2)
        print('正在统计')
        self.test_result = {}
        machine_list2 = list(dict.fromkeys(self.machine_list))
        test_item = ['EVM', 'MASK']
        self.meassage = {}
        xlsx_path = self.create_xlsx()
        one_record = 0
        for machine in machine_list2:
            self.meassage[machine] = {}
            self.test_result[machine] = {}
            self.test_result[machine]['uut'] = {}
            self.test_result[machine]['num'] = self.find_machine_num_and_uut(machine=machine)
            for te_item in test_item:
                self.creat_dict(item=te_item, machine=machine)
            for te_item in test_item:
                self.find_result(item=te_item, machine=machine)
            for uut in self.test_result[machine]['uut']:
                print_massage = ''
                for te_item in test_item:
                    ant_detail = ''
                    for ant in self.test_result[machine]['uut'][uut][te_item]['num_list']:
                        ant_detail = ant_detail + ant + f": {self.test_result[machine]['uut'][uut][te_item][ant]:02},     "
                    print_massage = print_massage + f"\n  {te_item}  {'总数:'}{self.test_result[machine]['uut'][uut][te_item]['amount']:02}{',     '}{ant_detail}"
                self.meassage[machine][uut] = print_massage
            book = openpyxl.load_workbook(xlsx_path)
            sheet = book['Sheet']
            sheet.cell(1, 1).value = '服务器'
            sheet.cell(1, 2).value = 'UUT'
            sheet.cell(1, 3).value = 'EVM'
            sheet.cell(1, 4).value = 'MASK'
            # len_uut = len(self.test_result[machine]['uut'])
            for gets_key in self.meassage[machine]:
                ff = 2 + one_record
                one_record += 1
                sheet.cell(ff, 1).value = machine
                sheet.cell(ff, 2).value = gets_key
                sheet.cell(ff, 3).value = self.meassage[machine][gets_key].split('\n')[1]
                sheet.cell(ff, 4).value = self.meassage[machine][gets_key].split('\n')[2]
            book.save(xlsx_path)
            book.close()

    def read_xslx_to_obj(self, path=None, sheet_name=None):
        book = openpyxl.load_workbook(f"{self.file_path}\\{path}")
        sheet = book[sheet_name]
        objects_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            obj = SnObject(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
                           row[11], row[12])
            print(obj.test)
            objects_list.append(obj)
        return objects_list


j = MainTool()
while True:
    chose_num = input('请选择 1:下载log\t2:查看测试结果\t3:log內容查詢\t4:统计 fail item\n')
    if chose_num == '1':
        break
    elif chose_num == '2':
        break
    elif chose_num == '3':
        break
    elif chose_num == '4':
        break
    else:
        continue  # 重新选择功能

if chose_num == '1':
    j.down_log()
elif chose_num == '2':
    j.check_test_result()
elif chose_num == '3':
    j.regular_text()
elif chose_num == '4':
    j.count_fail()

is_done = input('按任意按鍵結束')
