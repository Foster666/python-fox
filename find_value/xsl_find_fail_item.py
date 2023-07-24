import os
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
from matplotlib import pyplot as plt


class SnObject:

    def __init__(self, recttime, machine, container, tottime, testtime, sernum, uuttype, area, result, test,
                 parentsernum, username, mode):
        self.recttime = str(recttime)[:10]
        self.machine = machine
        self.container = container
        self.tottime = tottime
        self.testtime = testtime
        self.sernum = sernum
        self.uuttype = uuttype
        self.area = area
        self.result = result
        self.test = test
        self.parentsernum = parentsernum
        self.username = username
        self.mode = mode


class Find_fail_item:
    def __init__(self):
        self.file_path = r'F:\test_tool\dist\test_log'
        self.xls_name = 'test_build'  # xls 名称（可修修改）

    def read_xslx_to_obj(self, path=None, sheet_name=None):
        book = openpyxl.load_workbook(f"{self.file_path}\\{path}")
        sheet = book[sheet_name]
        objects_list = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            obj = SnObject(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10],
                           row[11], row[12])
            objects_list.append(obj)
        return objects_list

    def create_xlsx(self, xlsx_name=None):
        if xlsx_name:
            self.xl_name = xlsx_name + '.xlsx'
        else:
            self.xl_name = 'test_build.xlsx'
        try:
            try:
                os.remove(self.file_path + '\\' + self.xl_name)
                print('刪除文件成功,重新創建')
            except:
                pass
            self.book = openpyxl.Workbook()
            self.sheet = self.book['Sheet']
            col_1 = self.sheet.column_dimensions[get_column_letter(1)]
            col_1.width = 20
            print('文件創建成功')
        except:
            print(f'{xlsx_name}文件創建失敗')
        return self.file_path + '\\' + self.xl_name

    def crest_base_xslx(self, fail_item):
        for j_num in range(0, 10, 1):
            self.sheet.cell(j_num + 2, 1).value = f"5月{j_num + 2}日"
        for i_num in range(0, len(fail_item), 1):
            self.sheet.cell(1, i_num + 2).value = fail_item[i_num]
        self.sheet.cell(1, len(fail_item) + 2).value = 'all'
        self.sheet.cell(1, len(fail_item) + 3).value = 'total fail'

    def crest_analyze_xslx(self, fail_item):
        self.sheet.cell(1, 1).value = 'station'
        for i in range(2, 6, 1):
            self.sheet.cell(i, 1).value = 'PCBPM2'
        self.sheet.cell(2, 2).value = 'Tested qty'
        self.sheet.cell(3, 2).value = '1st passed qty'
        self.sheet.cell(4, 2).value = '1st failed qty'
        self.sheet.cell(5, 2).value = '1st pass yield'
        self.sheet.cell(1, 2).value = 'Date'

        self.sheet.cell(7, 1).value = 'cell number'
        self.sheet.cell(7, 2).value = 'total'
        self.sheet.cell(7, 3).value = 'FPY'
        self.sheet.cell(7, 4).value = 'top fail item&action'
        self.sheet.cell(8, 1).value = 'UUT04'
        self.sheet.cell(9, 1).value = 'UUT05'
        self.sheet.cell(10, 1).value = 'UUT06'
        self.sheet.cell(11, 1).value = 'UUT07'

    def draw_z(self, x_list=None, y_list=None, x_label=None, y_label=None, title=None, y_limit=None, png_name=None):
        if not y_limit:
            y_limit = 50
        y_float = [float(i) for i in y_list]
        fig, ax = plt.subplots()
        ax.plot(x_list, y_float, marker='o', color='red', linewidth=2)
        ax.set_xlabel(x_label, fontsize=12)
        ax.set_ylabel(y_label, fontsize=12)
        ax.set_title(title, fontsize=14)
        ax.set_ylim(bottom=y_limit, top=100)
        ax.tick_params(axis='both', which='major', labelsize=10)
        ax.grid(True, axis='y', linestyle='--', alpha=0.7)
        # plt.show()
        plt.savefig(f'{self.file_path}\\{png_name}.png', bbox_inches='tight')
        print(f'已保存{png_name}')

    def draw_zhu(self, x_list=None, y_list=None, x_label=None, y_label=None, title=None, png_name=None):
        fig, ax = plt.subplots()
        # 绘制柱状图
        ax.bar(x_list, y_list)
        # 设置横轴标签
        ax.set_xlabel(x_label)
        # 设置纵轴标签
        ax.set_ylabel(y_label)
        # 设置标题
        ax.set_title(title)
        # 自适应调整横轴标签旋转角度
        plt.xticks(rotation=10, ha='right')
        # 显示图形
        for i in range(len(x_list)):
            plt.text(i, y_list[i] + 0.2, str(y_list[i]), ha="center")
        # plt.show()
        plt.savefig(f'{self.file_path}\\{png_name}.png', bbox_inches='tight')
        print(f'已保存{png_name}')

    def count_fail(self):
        self.create_xlsx()
        fail_item = ['25_TX_CAL_A0', 'PCBPM2_RUN_TEST_FLOW', '180_TX_5180', '6_TX_CAL_A0', '12_TX_CAL', '281_RX_VERIFY',
                     'PDA_', '18_TX', 'PING_SERVER', 'POE_AT', '5_XTAL']
        self.crest_base_xslx(fail_item)
        for i in range(2, 12, 1):
            print(i)
            self.objects_list = self.read_xslx_to_obj(path=f"5-{i}.xlsx", sheet_name='Sheet1')
            self.objects_list_fail = []
            for sn_object_all in self.objects_list:
                if sn_object_all.result == 'F':
                    self.objects_list_fail.append(sn_object_all)
            total_quantity = len(self.objects_list_fail)
            total_percent = 0
            for f_num in range(0, len(fail_item), 1):
                one_quantity = 0
                for sn_obj_fail in self.objects_list_fail:
                    try:
                        sn_obj_fail.test.index(fail_item[f_num])
                        one_quantity += 1
                    except:
                        pass
                persent_item = one_quantity * 100 / total_quantity
                total_percent += round(persent_item, 2)
                one_percent = round(persent_item, 2)
                self.sheet.cell(i, 2 + f_num).value = one_percent
                self.sheet.cell(i, len(fail_item) + 2).value = f"{total_percent:.2f}%"
                self.sheet.cell(i, len(fail_item) + 3).value = total_quantity
        self.book.save(self.file_path + '\\' + self.xl_name)
        self.book.close()

    def count_item(self, fail_item, top_num):
        item_dict = {}
        for item in fail_item:
            item_dict[item] = 0
        for list_item in self.list_fail:
            try:
                item_dict[list_item.test] += 1
            except:
                pass
        sorted_items = sorted(item_dict.items(), key=lambda x: x[1], reverse=True)
        top_items = dict(sorted_items[:top_num])
        return top_items

    def analyze(self):
        self.create_xlsx()
        fail_item = ['12_TX_CAL_A1', '18_TX_CAL_B', '180_TX_5180_ANT1__MASK', '25_TX_CAL_A0',
                     '6_TX_CAL_A0', 'PCBPM2_RUN_TEST_FLOW', 'PDA_ALT-A/B_MDI', 'PING_SERVER_1000M',
                     'POE_AT_NOT_FOUND']
        self.crest_analyze_xslx(fail_item)
        self.objects_list = self.read_xslx_to_obj(path=f"5-month.xlsx", sheet_name='Sheet1')
        container_list = ['PCBPM2:UUT04', 'PCBPM2:UUT05', 'PCBPM2:UUT06', 'PCBPM2:UUT07']
        uut_4_x = []
        uut_4_y = []
        uut_5_x = []
        uut_5_y = []
        uut_6_x = []
        uut_6_y = []
        uut_7_x = []
        uut_7_y = []
        for date_num in range(2, 32, 1):
            if date_num == 14:
                continue
            self.date_list = []
            for sn_date_list in self.objects_list:
                if sn_date_list.recttime == f'2023-05-{date_num:02}':
                    self.date_list.append(sn_date_list)
            # print(f"2023-05-{date_num}当天测试数量：{len(self.date_list)}")
            for con_num in range(len(container_list)):
                self.list_fail = []
                self.container_list = []
                p_list = []
                f_list = []
                for sn_p in self.date_list:
                    if sn_p.result == 'P':
                        p_list.append(sn_p)
                for sn_f in self.date_list:
                    if sn_f.result == 'F':
                        f_list.append(sn_f)
                for sn_container_list in self.date_list:
                    if sn_container_list.container == container_list[con_num]:
                        self.container_list.append(sn_container_list)
                # print(f"{container_list[con_num]},共测试{len(self.container_list)}")
                for list_fail in self.container_list:
                    if list_fail.result == 'F':
                        self.list_fail.append(list_fail)
                # print(f"{container_list[con_num]}一共fail  {len(self.list_fail)}")
                # for kk in self.list_fail:
                #     print(kk.test)
                top_items = self.count_item(fail_item=fail_item, top_num=3)
                # print(top_items)
                self.sheet.cell(8 + con_num, 2 + (date_num - 2) * 3).value = len(self.container_list)
                self.sheet.cell(8 + con_num, 3 + (
                        date_num - 2) * 3).value = f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}%"
                self.sheet.cell(13, 3 + (date_num - 2) * 1).value = f"5-{date_num}"
                self.sheet.cell(14 + con_num, 3 + (
                        date_num - 2) * 1).value = f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}%"
                self.sheet.cell(8 + con_num, 4 + (date_num - 2) * 3).value = str(top_items)
                self.sheet.cell(1, 3 + date_num).value = f'05-{date_num:02}'
                self.sheet.cell(2, 3 + date_num).value = len(self.date_list)
                self.sheet.cell(3, 3 + date_num).value = len(p_list)
                self.sheet.cell(4, 3 + date_num).value = len(f_list)
                self.sheet.cell(5, 3 + date_num).value = str(f"{100 * len(p_list) / len(self.date_list):.2f}%")
                # print("******************************************************")
                if container_list[con_num] == "PCBPM2:UUT04":
                    uut_4_x.append(f"{date_num}")
                    uut_4_y.append(
                        f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}")
                elif container_list[con_num] == "PCBPM2:UUT05":
                    uut_5_x.append(f"{date_num}")
                    uut_5_y.append(
                        f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}")
                elif container_list[con_num] == "PCBPM2:UUT06":
                    uut_6_x.append(f"{date_num}")
                    uut_6_y.append(
                        f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}")
                elif container_list[con_num] == "PCBPM2:UUT07":
                    uut_7_x.append(f"{date_num}")
                    uut_7_y.append(
                        f"{100 * (len(self.container_list) - len(self.list_fail)) / len(self.container_list):.2f}")
        self.draw_z(x_list=uut_4_x, y_list=uut_4_y, x_label='日期', y_label='良率', title='uut04(5.2-31)', y_limit=30,
                    png_name='uut04')
        self.draw_z(x_list=uut_5_x, y_list=uut_5_y, x_label='日期', y_label='良率', title='uut05(5.2-31)', y_limit=30,
                    png_name='uut05')
        self.draw_z(x_list=uut_6_x, y_list=uut_6_y, x_label='日期', y_label='良率', title='uut06(5.2-31)', y_limit=30,
                    png_name='uut06')
        self.draw_z(x_list=uut_7_x, y_list=uut_7_y, x_label='日期', y_label='良率', title='uut07(5.2-31)', y_limit=30,
                    png_name='uut07')

        self.book.save(self.file_path + '\\' + self.xl_name)
        self.book.close()

    def average_value(self):
        book = openpyxl.load_workbook(f"{self.file_path}\\FOC271924UN_BPM2_CPK.xlsx")
        sheet = book["Sheet"]
        for i in range(0, 6, 1):
            num_18 = 0 + 8 * i
            chu_num = float(4)
            AA_UP = 0
            BB_UP = 0
            CC_UP = 0
            DD_UP = 0
            AA_LO = 0
            BB_LO = 0
            CC_LO = 0
            DD_LO = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # print(float(row[1+num_18]))
                AA_UP += float(row[1 + num_18])
                BB_UP += float(row[2 + num_18])
                CC_UP += float(row[3 + num_18])
                DD_UP += float(row[4 + num_18])
                AA_LO += float(row[5 + num_18])
                BB_LO += float(row[6 + num_18])
                CC_LO += float(row[7 + num_18])
                DD_LO += float(row[8 + num_18])
            # print(f"AA_UP{AA_UP/1365},BB_UP{BB_UP/1365},CC_UP{CC_UP/1365},DD_UP{DD_UP/1365},AA_LO{AA_LO/1365},BB_LO{BB_LO/1365},CC_LO{CC_LO/1365},DD_LO{DD_LO/1365}")
            # print(f"AA_UP {AA_UP / chu_num:.2f},BB_UP {BB_UP / chu_num:.2f},CC_UP {CC_UP / chu_num:.2f},DD_UP {DD_UP / chu_num:.2f},AA_LO {AA_LO / chu_num:.2f},BB_LO {BB_LO / chu_num:.2f},CC_LO {CC_LO / chu_num:.2f},DD_LO {DD_LO / chu_num:.2f}")
            print(f"18{i}.")
            print(f"MARGIN_DB_UP_A :       {AA_UP / chu_num:.2f}   dB")
            print(f"MARGIN_DB_UP_B :       {BB_UP / chu_num:.2f}   dB")
            print(f"MARGIN_DB_UP_C :       {CC_UP / chu_num:.2f}   dB")
            print(f"MARGIN_DB_UP_D :       {DD_UP / chu_num:.2f}   dB")
            print(f"MARGIN_DB_LO_A :       {AA_LO / chu_num:.2f}   dB")
            print(f"MARGIN_DB_LO_B :       {BB_LO / chu_num:.2f}   dB")
            print(f"MARGIN_DB_LO_C :       {CC_LO / chu_num:.2f}   dB")
            print(f"MARGIN_DB_LO_D :       {DD_LO / chu_num:.2f}   dB")
            print('')

    # def count_qty(self):
    def analyze_manual(self):
        fail_item = ['12_TX_CAL_A1', '18_TX_CAL_B', '180_TX_5180_ANT1__MASK', '25_TX_CAL_A0',
                     '6_TX_CAL_A0', 'PCBPM2_RUN_TEST_FLOW', 'PDA_ALT-A/B_MDI', 'PING_SERVER_1000M',
                     'POE_AT_NOT_FOUND']
        x_2_7 = ['6_TX_CAL_A0', '180_TX_5180_ANT1__MASK', '25_TX_CAL_A0', 'PCBPM2_RUN_TEST_FLOW', '12_TX_CAL_A1']
        y_2_7 = [83, 34, 25, 21, 10]
        x_8_14 = ['180_TX_5180_ANT1__MASK', '6_TX_CAL_A0', 'PDA_ALT-A/B_MDI', '25_TX_CAL_A0', '12_TX_CAL_A1']
        y_8_14 = [46, 39, 37, 26, 23]
        x_15_21 = ['180_TX_5180_ANT1__MASK', '6_TX_CAL_A0', '12_TX_CAL_A1', '25_TX_CAL_A0', 'PDA_ALT-A/B_MDI']
        y_15_21 = [100, 43, 28, 27, 15]
        x_22_29 = ['180_TX_5180_ANT1__MASK', '6_TX_CAL_A0', '12_TX_CAL_A1', '18_TX_CAL_B', '33_TX_5180_ANT1']
        y_22_29 = [150, 32, 16, 10, 7]
        self.draw_zhu(x_list=x_2_7, y_list=y_2_7, x_label='', y_label='数量', title='TOP5不良(5.2-5.7)',
                      png_name='5_2-7_TOP_fail')
        self.draw_zhu(x_list=x_8_14, y_list=y_8_14, x_label='', y_label='数量', title='TOP5不良(5.8-5.14)',
                      png_name='5_8-14_TOP_fail')
        self.draw_zhu(x_list=x_15_21, y_list=y_15_21, x_label='', y_label='数量', title='TOP5不良(5.15-5.21)',
                      png_name='5_15-21_TOP_fail')
        self.draw_zhu(x_list=x_22_29, y_list=y_22_29, x_label='', y_label='数量', title='TOP5不良(5.22-5.29)',
                      png_name='5_22-29_TOP_fail')


j = Find_fail_item()
j.average_value()
# yy_list = [9,7,1,0,1,1,1,0,0,0,0,0,0,0,0,0,0,0,0]
# xx_lixt = ['5-11','12','13','14','15','16','17','18','19','20','21','22','23','24','25','26','27','28','5-29']
# j.draw_zhu(x_list=xx_lixt, y_list=yy_list, x_label='日期', y_label='数量', title='pda相关不良(5.22-5.29)', png_name='pda_fail')
